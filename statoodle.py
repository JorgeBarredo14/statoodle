import tkinter as tk
import tkinter.ttk as ttk
from tkinter.filedialog import *
import tkinter.filedialog as filedialog
from tkinter import StringVar
from tkinter import LabelFrame
from tkinter import Label
from tkinter import Button
from tkinter import messagebox

from fpdf import FPDF
import openpyxl
from pathlib import Path
import matplotlib as mpl
from matplotlib import pyplot as plt

import pandas as pd
import numpy as np
import math
import os
from ttkwidgets.autocomplete import AutocompleteEntry
from fpdf import FPDF
import pathlib
from datetime import datetime, date
import re

global rootName
global filename
global whole_marks
global efficiency

plt.ioff()

root = tk.Tk()
menubar = tk.Menu(root)
root.config(menu= menubar)
root.title("Statoodle")
subMenu = tk.Menu(menubar, tearoff=0)
root.geometry("500x580")
rootName = "Course"
filename = "Unselected folder"
var = StringVar()

def check_columns(x,to_check):
    if not set(to_check).issubset(set(x.columns)):
       return f"{', '.join(set(to_check).difference(x.columns))} are not available in the dataframe."
    return "All columns are available in the dataframe."

def ScoreMark(score):
    if ( score >= 9 ):
        mark = 'SOBRESALIENTE (SB)'
    elif ( score >= 7):
        mark = 'NOTABLE (NT)'
    elif ( score >= 5):
        mark = 'APROBADO (AP)'
    else: 
        mark = 'SUSPENSO (SS)'
    return mark

def SingleQuizReport(path, shortname, filename, pdf, names, pb): 
        
    try:
        data = pd.read_excel(filename)
        n_students=np.shape(data)[0]-1
        data_cols = data.columns.values.tolist()
        data_cols[7]=data_cols[7].replace(",", ".")
        res=re.search('Calificación/(.*).', str(data_cols[7]))
        max_mark=float(res.group(1))
        mpl.rcParams["font.size"] = 10

         # Calculation of average exam time
        
        print(n_students)
        total_time = 0
        students_required_time=np.zeros(n_students)
                
        for i in range(n_students):
            #print(i)
            res=re.findall(r"(\d+) hora", data.loc[i,data_cols[6]])
            if res:
                students_required_time[i] =  students_required_time[i]+float(res[0])*60
            #print(students_required_time[i])
            res=re.findall(r"(\d+) minuto", data.loc[i,data_cols[6]])
            if res:
                students_required_time[i] =  students_required_time[i]+float(res[0])
            #print(students_required_time[i])
            res=re.findall(r"(\d+) segundo", data.loc[i,data_cols[6]])
            if res:
                students_required_time[i] =  students_required_time[i]+float(res[0])/60
            #print(students_required_time[i])
       
        average_time = np.round(np.mean(students_required_time), decimals=2)
        print("Average required time in the exam:", average_time, "minutes")
        pb['value']=20
        root.update_idletasks()

        # Print table with points/min
        login_data=pd.read_excel(filename) 
        login_data=login_data[:-1]
        backup_ld = login_data.copy()
        backup_ld = backup_ld.replace("NP", "0")
        backup_ld = backup_ld.replace("-", "0")
        backup_ld = backup_ld.stack().str.replace(',','.').unstack()
        cols=[i for i in backup_ld.columns if i not in ["Apellido(s)", "Nombre", "Dirección de correo", "Estado", "Comenzado el", "Finalizado", "Tiempo requerido"]]
        for col in cols:
            backup_ld[col]= pd.to_numeric(backup_ld[col], downcast="float")/max_mark*10
        points_per_min=[]
        points_per_min=backup_ld[login_data.columns.values.tolist()[7]]/students_required_time
        points_per_min=np.round(points_per_min, decimals=2)
        plt.figure(figsize = (25, 10), dpi = 80)
        plt.xticks(range(len(names)),names,rotation=90)
        plt.bar(names, points_per_min)
        plt.title("Student's efficiency")
        plt.xlabel("Student", fontsize=14)
        plt.ylabel("Efficiency (points per minute)", fontsize=14)
        for index, value in enumerate(points_per_min):
            plt.text(index,value, str(np.round(value,decimals=2)), fontsize=14)
        plt.savefig(path + '/quiz_efficiency.png')
        plt.show()
        pb['value']=30
        root.update_idletasks()

        # Get stats of average marks
        avg_mark = backup_ld[login_data.columns.values.tolist()[7]].mean()
        print("Average mark in the exam: ", avg_mark)

        califications_conversion = []
        califications_conversion = backup_ld[login_data.columns.values.tolist()[7]].apply(lambda s: ScoreMark(int(s)))
        califications_conversion.value_counts(sort = True).plot.pie(autopct='%1.2f%%', title="Students' marks",shadow=True)
        plt.ylabel("")
        plt.savefig(path + '/quiz_students_marks.png')

        # Print marks bar chart
        fig = plt.figure(figsize = (5,6))
        backup_ld[login_data.columns.values.tolist()[7]].astype(int).value_counts(sort=False).plot.bar(rot=0)
        plt.xlabel('Integer value of the mark')
        plt.ylabel('Distribution of marks')
        plt.title("Counts of intervals for each student's mark")
        plt.grid()
        plt.savefig(path + '/quiz_calification_barchart.png')
        plt.show()

        # Print distribution of marks for each question
        question_grades=backup_ld.drop(columns=["Estado","Comenzado el","Finalizado","Tiempo requerido"])
        fig = plt.figure(figsize = (15,10))
        ax = fig.gca()
        question_grades.hist(ax=ax)
        plt.suptitle("Distribution of marks for each question")
        plt.xlabel('Integer value of the mark')
        plt.ylabel('Distribution of marks')
        plt.savefig(path + '/questions_calification_barchart.png')
        plt.show()

        # Get most difficult questions
        n_questions=np.shape(backup_ld)[1]-8
        print(n_questions)
        avg_questions_marks=np.zeros(n_questions)
        for i in range(n_questions):
            avg_questions_marks[i] = backup_ld.iloc[:,i+8].mean()
            print("Question", i+1, avg_questions_marks[i])
        print("Avg_questions_marks", avg_questions_marks)
        n_difficult_questions = math.floor(0.2*n_questions)
        most_difficult_questions = sorted(range(len(avg_questions_marks)), key = lambda sub: avg_questions_marks[sub])[:n_difficult_questions]
        for i in range(len(most_difficult_questions)):
            most_difficult_questions[i] = most_difficult_questions[i] + 1
        #print("Most difficult questions have been", " " . join(str(a) for a in most_difficult_questions))

        pb['value']=40
        root.update_idletasks()

        # Calculation of IRT
        avg_marks=backup_ld[data.columns.values.tolist()[7]].to_numpy()/10/max_mark*10
        avg_marks=np.where(avg_marks==1,0.999,avg_marks)

        theta = np.log(avg_marks/(1-avg_marks)) # level of the student
        #print(theta)

        questions_means=backup_ld.mean()[2:].to_numpy()
        questions_means=np.where(questions_means==1,0.999,questions_means)
        print("Questions_means", questions_means)
        #print(backup_ld)
        beta = np.log((1-questions_means)/questions_means) # difficulty of the question
        for it in range(len(beta)):
            if(beta[it] > 3):
                beta[it] = 3
            if(beta[it] < -3):
                beta[it] = -3
            if(beta[it] == math.inf):
                beta[it] = 3
            if(beta[it] == -math.inf):
                beta[it] == -3
        #print("Beta", beta)

        probability=[]

        for i in range(len(theta)):
            probability.append([])
            for j in range(len(beta)):
                probability[i].append(np.exp(theta[i]-beta[j])/(1+np.exp(theta[i]-beta[j]))) # probability of answering correctly
        #print(probability)

        # Print normalized beta
        division=np.max(beta) - np.min(beta)
        if division==0:
            division=0.0001
        norm_beta = 10*(beta - np.min(beta)) / division
        for it in range(len(norm_beta)):
            if norm_beta[it] == math.inf:
                norm_beta[it]=3
        norm_beta = np.round(norm_beta, decimals=2)
        print(norm_beta)
        print(beta)
        plt.figure(figsize = (25, 10), dpi = 80)
        plt.xticks(range(0,n_questions),range(1,n_questions+1))
        mpl.rcParams["font.size"] = 24
        for index, value in enumerate(norm_beta):
            plt.text(index,value, str(value))
        plt.xlabel('Question', fontsize=24)
        plt.ylabel('Difficulty', fontsize=24)
        print(n_questions,np.shape(norm_beta))
        plt.bar(range(n_questions),norm_beta)
        plt.savefig(path + '/quiz_difficulty.png')
        plt.show()

        # Print probabilities
        probability=np.array(probability).reshape(n_students,n_questions)
        #print(probability)
        plt.figure(figsize = (50, 30), dpi = 80)
        legend_question = []
        #legend_question = "Question" + range(0,n_questions)
        legend_question = ["{}{}".format("Q", i) for i in range(1, n_questions+1)]
        plt.plot(probability, linewidth=3)
        plt.legend(legend_question)
        plt.grid()
        plt.xticks(range(0,n_students),names,rotation=90)
        plt.xlabel('Student', fontsize=24)
        plt.ylabel('Probability of answering correctly', fontsize=24)
        plt.scatter(range(0,n_students),probability.mean(axis=1), label="Mean probability for each student", linewidth=10)
        plt.plot(range(0,n_students),probability.mean(axis=1), linewidth=5)
        plt.legend()
        plt.savefig(path + '/quiz_IRT.png')
        plt.show()

        # Print IRT curves
        mpl.rcParams["font.size"] = 10
        probability=[]
        theta = np.linspace(-3, 3, num=100)
        mean_beta = np.mean(beta)
        for i in range(len(beta)):
           probability.append([])
           for j in range(len(theta)):
               probability[i].append(np.exp(theta[j]-beta[i])/(1+np.exp(theta[j]-beta[i]))) # probability of answering correctly
        #print(probability)

        # Print probabilities
        probability=np.array(probability).reshape(len(beta),len(theta))
        #print(probability)

        # Create directory for ICCs
        ICCdirName = 'ICC_' + os.path.basename(filename)
        try:
           # Create target Directory
           os.mkdir(path + "/" + ICCdirName)
           print("Directory " , ICCdirName ,  " Created ") 
        except FileExistsError:
           print("Directory " , ICCdirName ,  " already exists")

        # Create ICCs    
        #slope_arr=np.zeros(shape=(np.shape(probability)[0],2))
        #slope=np.zeros(shape=(np.shape(probability)[0],1))
        for k in range(0,np.shape(probability)[0]):
            plt.figure(figsize = (5, 4), dpi = 80)
            plt.plot(theta,probability[k,:])
            #slope_arr[k][0]=np.interp(0.49,probability[k,:],theta)
            #slope_arr[k][1]=np.interp(0.51,probability[k,:],theta)
            #slope[k]=(0.51-0.49)/(slope_arr[k][1]-slope_arr[k][0])
            #print(slope)
            plt.grid()
            #plt.xticks(range(0,n_students),names,rotation=90)
            plt.xlabel('\u03B8', fontsize=12)
            plt.ylabel('P(\u03B8)', fontsize=12)
            #plt.scatter(theta,probability.mean(axis=1), label="Mean probability for each student")
            #plt.plot(range(0,n_students),probability.mean(axis=1))
            #plt.legend()
            plt.title("ICC for Question {}" .format(k+1))
            plt.savefig(path + "/" + ICCdirName + "/ICC_{:03}.png".format(k+1))
            plt.show()
        #print(slope_arr)
        #print(slope)
        pb['value']=50
        root.update_idletasks()

        pdf.add_page()
        pdf.set_xy(0, 0)
        pdf.set_font('times', 'B', 24)
        pdf.cell(60)
        pdf.cell(90, 20, "Exam report: " + str(os.path.basename(filename)), 0, 2, 'C')
        pdf.cell(-40)
        pdf.set_font('times', '', 12)
        pdf.cell(100, 25, txt="The average time required by students for completing the exam was " + str(average_time) + " minutes, and")
        pdf.cell(-100)
        pdf.cell(0, 40, txt="the average mark was " + str(np.round(avg_mark, decimals=2)) + ". Attending to the results, the most difficult questions for the")
        pdf.cell(-180)
        pdf.cell(0, 55, txt= "students were: " + str(" " . join(str(a) for a in most_difficult_questions)) + ".")
        pdf.cell(-180)
        pdf.set_font('times', 'B', 15)
        pdf.cell(0, 80, "Stats related to the students marks", 0, 2, 'L')
        pdf.set_font('times', '', 12)
        pdf.cell(100, -60, txt= "This exam reported the following statistics. The pie chart represents the percentage of marks in integer")
        pdf.cell(-100)
        pdf.cell(100, -45, txt= "intervals, and the right histogram shows the quantity of students that have their mark in that interval.")
        pdf.cell(-100)
        pdf.cell(100, -30, txt= "Last but not least, the last group of histograms shows the balance between wrong and correct answers.")
        pdf.image(path + '/quiz_students_marks.png', x = 10, y = 90, w = 90, h = 70, type = '', link = '')
        pdf.image(path + '/quiz_calification_barchart.png', x = 100, y = 90, w = 60, h = 65, type = '', link = '')
        pdf.image(path + '/questions_calification_barchart.png', x = 10, y = 160, w = 160, h = 140, type = '', link = '')

        pdf.add_page()
        pdf.set_xy(0, 0)
        pdf.set_font('times', 'B', 24)
        pdf.cell(60)
        pdf.cell(90, 20, "Stats related to the exam", 0, 2, 'L')
        pdf.cell(-40)
        pdf.set_font('times', '', 12)
        pdf.cell(100, 25, txt="In order to fulfil these exam stats, some algorithms have been executed, such as Item Response Theory (IRT).")
        pdf.cell(-100)
        pdf.cell(0, 40, txt="It resulted on a scale that shows the difficulty of the questions, attending to a customized and normalized")
        pdf.cell(-180)
        pdf.cell(0, 55, txt= "measurement. The probability of answering correctly each question for every student is also calculated.")
        pdf.image(path + '/quiz_difficulty.png', x = 0, y = 50, w = 200, h = 100, type = '', link = '')
        pdf.cell(-60)

        pdf.add_page(orientation="L")
        pdf.set_xy(0,0)
        pdf.set_font('times', 'B', 18)
        pdf.cell(100)
        pdf.cell(90, 20, "Student's resolution ability", 0, 2, 'C')
        pdf.image(path + '/quiz_IRT.png', x = -15, y = 15, w = 320, h = 180, type = '', link = '')

        pdf.add_page(orientation="L")
        pdf.set_xy(0,0)
        pdf.cell(104)
        pdf.set_font('times', 'B', 18)
        pdf.cell(90, 20, "Student's resolution efficiency", 0, 2, 'C')
        pdf.image(path + '/quiz_efficiency.png', x = -15, y = 15, w = 320, h = 190, type = '', link = '')
    except AttributeError:
        pass
    
def course_pdf(path, pdf, pb):
    pdf.add_page()
    pdf.set_xy(0, 0)
    pdf.set_font('times', 'B', 24)
    pdf.cell(60)
    pdf.cell(90, 20, "Course report", 0, 2, 'C')
    pdf.cell(-40)
    pdf.set_font('times', 'B', 15)
    pdf.cell(0, 80, "Stats related to the students marks", 0, 2, 'L')
    pdf.set_font('times', '', 12)
    pdf.cell(100, -60, txt= "This exam reported the following statistics. The pie chart represents the percentage of marks in integer")
    pdf.cell(-100)
    pdf.cell(100, -45, txt= "intervals, and the right histogram shows the quantity of students that have their mark in that interval.")
    pdf.cell(-100)
    pdf.cell(100, -30, txt= "Last but not least, the last group of histograms shows the balance between wrong and correct answers.")
    pdf.image(path + "whole_course_piechart.png", x = 10, y = 90, w = 90, h = 70, type = '', link = '')
    pdf.image(path + "whole_course_barchart.png", x = 100, y = 90, w = 60, h = 65, type = '', link = '')
    pdf.image(path + "whole_course_exams_barchart.png", x = 10, y = 160, w = 160, h = 140, type = '', link = '')

    pdf.add_page()
    pdf.set_xy(0, 0)
    pdf.set_font('times', 'B', 24)
    pdf.cell(60)
    pdf.cell(90, 20, "Stats related to the course", 0, 2, 'L')
    pdf.cell(-40)
    pdf.set_font('times', '', 12)
    pdf.cell(100, 25, txt="In order to fulfil these exam stats, some algorithms have been executed, such as Item Response Theory (IRT).")
    pdf.cell(-100)
    pdf.cell(0, 40, txt="It resulted on a scale that shows the difficulty of the questions, attending to a customized and normalized")
    pdf.cell(-180)
    pdf.cell(0, 55, txt= "measurement. The probability of answering correctly each question for every student is also calculated.")
    pdf.image(path + '/dif_questions.png', x = 0, y = 50, w = 200, h = 100, type = '', link = '')
    pdf.cell(-60)

    pdf.add_page(orientation="L")
    pdf.set_xy(0,0)
    pdf.set_font('times', 'B', 18)
    pdf.cell(100)
    pdf.cell(90, 20, "Student's resolution ability", 0, 2, 'C')
    pdf.image(path + "course_probability_IRT.png", x = -15, y = 15, w = 320, h = 180, type = '', link = '')

    pdf.add_page(orientation="L")
    pdf.set_xy(0,0)
    pdf.cell(104)
    pdf.set_font('times', 'B', 18)
    pdf.cell(90, 20, "Student's course resolution efficiency", 0, 2, 'C')
    pdf.image(path + "course_efficiency.png", x = -15, y = 15, w = 320, h = 190, type = '', link = '')
    
    pdf.add_page(orientation="L")
    pdf.set_xy(0,0)
    pdf.set_font('times', 'B', 18)
    pdf.cell(100)
    pdf.image(path + 'course_marks_normal_distribution.png', x = -15, y = 15, w = 320, h = 180, type = '', link = '')
      
def openSingleQuizReport(): 
    entered=0
    root.update()
    filename = filedialog.askopenfilename(title="Choose a Quiz Report",filetypes=(("Excel Files (.xlsx)","*.xlsx"),("All Files","*.*")))
    root.update()
    qr_cols=['Apellido(s)', 'Nombre', 'Dirección de correo', 'Estado', 'Comenzado el', 'Finalizado', 'Tiempo requerido']
    pb = ttk.Progressbar(root, length=100, mode='determinate')
    pb.grid(column=0, padx=200, sticky="w")
    lab = tk.Label(root, text='Calculating...')
    lab.grid(column=0, padx=210, sticky="w")
    pb['value']=0
    root.update_idletasks()
    try:
        entered=1    
        pdf = FPDF()
        shortname, sep, tail = (os.path.basename(filename)).partition('.')
        dirName = 'Quiz_' + shortname
        path = dirName
        data = pd.ExcelFile(filename).parse()
        data.drop(data.tail(1).index,inplace=True)
        columns_ok=check_columns(data, qr_cols)
        print(columns_ok)
        if columns_ok == "All columns are available in the dataframe.":
            try:
                os.mkdir(path)
                print("Directory " + path + " created")
            except FileExistsError:
                print("Directory " + path + " already exists")
            # Read names of students
            names=[]
            names = data[['Apellido(s)', 'Nombre']].apply(lambda x: ', '.join(x[x.notnull()]), axis = 1)
            pb['value']=10
            root.update_idletasks()
            SingleQuizReport(path,shortname,filename, pdf, names, pb)
            pb['value']=100
            lab.destroy()
            pb.grid_forget()
            root.update_idletasks()
            root.update()
            pdf.output(asksaveasfilename(title="Save output file", defaultextension=".pdf", filetypes=(("PDF File", "*.pdf"),("All Files", "*.*"))), "F")
            root.update()
        else:
            pb['value']=100
            lab.destroy()
            pb.grid_forget()
            root.update_idletasks()
            messagebox.showinfo("Warning", "File not analyzed: " + columns_ok) 
    except (FileNotFoundError,IndexError):
        pb['value']=100
        lab.destroy()
        pb.grid_forget()
        root.update_idletasks()
        messagebox.showinfo("Warning", "No file has been chosen!")
    except ValueError:
        pb['value']=100
        lab.destroy()
        pb.grid_forget()
        root.update_idletasks()
        messagebox.showinfo("Warning", "No Excel file has been chosen!")
    
def openMultipleQuizReport(): 
    entered=0
    root.update()
    filenames = filedialog.askopenfilenames(title="Choose multiple Quiz Reports",filetypes=(("Excel Files (.xlsx)","*.xlsx"),("All Files","*.*")))
    root.update()
    qr_cols=['Apellido(s)', 'Nombre', 'Dirección de correo', 'Estado', 'Comenzado el', 'Finalizado', 'Tiempo requerido', 'Calificación/10,00']
    try:
        pb = ttk.Progressbar(root, length=100, mode='determinate')
        pb.grid(column=0, padx=200, sticky="w")
        lab = tk.Label(root, text='Calculating...')
        lab.grid(column=0, padx=210, sticky="w")
        entered=1    
        pdf = FPDF()
        shortname, sep, tail = (os.path.basename(filenames[0])).partition('.')
        dirName = 'Quiz_' + shortname
        if len(filenames)>1:
            whole_marks=pd.DataFrame()
            whole_marks["Students"]=""
            efficiency=pd.DataFrame()
            efficiency["Students"]=""  
            for index in range(len(filenames)):
                root_ext = os.path.splitext(os.getcwd())
                parent_dir = str(pathlib.PureWindowsPath(root_ext[0]).as_posix()) + "/"
                parent_dir = root_ext[0] + "/"
                filename=filenames[index]
                shortname, sep, tail = (os.path.basename(filename)).partition('.')
                dirName = 'Quiz_' + shortname
                data = pd.ExcelFile(filename).parse()
                data.drop(data.tail(1).index,inplace=True)
                columns_ok=check_columns(data, qr_cols)
                print(columns_ok)
                
                if columns_ok == "All columns are available in the dataframe.":
                    # Create directory for each quiz
                    try:
                        os.mkdir("Course")
                        print("Directory 'Course' created")
                    except FileExistsError:
                        print("Directory 'Course' already exists")    
                    path = parent_dir + "Course/" + dirName
                    # Create directory for each quiz
                    try:
                        os.mkdir(path)
                        print("Directory" + path + "created")
                    except FileExistsError:
                        print("Directory" + path + "already exists")      
                    # Read names of students
                    names=[]
                    names = data[['Apellido(s)', 'Nombre']].apply(lambda x: ', '.join(x[x.notnull()]), axis = 1)
                    pb['value']=10
                    root.update_idletasks()
                    SingleQuizReport(path, shortname, filename, pdf, names, pb)
                    
                    pb['value']=70
                    root.update_idletasks()
                    
                    # Check that this exam is not in the marks column
                    if shortname not in whole_marks.columns:
                        whole_marks[shortname]=np.nan
                        efficiency[shortname]=np.nan
                    # Calculation of average exam time
                    data = pd.read_excel(filename)
                    data_cols = data.columns.values.tolist()
                    data_cols[7]=data_cols[7].replace(",", ".")
                    res=re.search('Calificación/(.*).', str(data_cols[7]))
                    max_mark=float(res.group(1))
                    mpl.rcParams["font.size"] = 10

                    n_students=np.shape(data)[0]-1
                    print(n_students)
                    total_time = 0
                    students_required_time=np.zeros(n_students)

                    for i in range(n_students):
                        #print(i)
                        res=re.findall(r"(\d+) hora", data.loc[i,data_cols[6]])
                        if res:
                            students_required_time[i] =  students_required_time[i]+float(res[0])*60
                        #print(students_required_time[i])
                        res=re.findall(r"(\d+) minuto", data.loc[i,data_cols[6]])
                        if res:
                            students_required_time[i] =  students_required_time[i]+float(res[0])
                        #print(students_required_time[i])
                        res=re.findall(r"(\d+) segundo", data.loc[i,data_cols[6]])
                        if res:
                            students_required_time[i] =  students_required_time[i]+float(res[0])/60
                        #print(students_required_time[i])
                    average_time = np.round(np.mean(students_required_time), decimals=2)

                    # Calculation of average exam time                    
                    login_data=pd.read_excel(filename) 
                    login_data=login_data[:-1]
                    data_cols = login_data.columns.values.tolist()
                    data_cols[7]=data_cols[7].replace(",", ".")
                    res=re.search('Calificación/(.*).', str(data_cols[7]))
                    max_mark=float(res.group(1))
                    backup_ld = login_data.copy()
                    backup_ld = backup_ld.replace("NP", "0")
                    backup_ld = backup_ld.replace("-", "0")
                    backup_ld = backup_ld.stack().str.replace(',','.').unstack()
                    cols=[i for i in backup_ld.columns if i not in ["Apellido(s)", "Nombre", "Estado", "Comenzado el", "Finalizado", "Tiempo requerido"]]
                    for col in cols:
                        backup_ld[col]= pd.to_numeric(backup_ld[col], downcast="float")/max_mark*10
                    points_per_min=[]
                    points_per_min=backup_ld[login_data.columns.values.tolist()[7]]/students_required_time
                    points_per_min=np.round(points_per_min, decimals=2)
                    for student in range(len(names)):
                                if ~(whole_marks['Students']==names[student]).any():
                                    whole_marks=whole_marks.append({'Students' : names[student]} , ignore_index=True)
                                    efficiency=efficiency.append({'Students' : names[student]} , ignore_index=True)
                                marks_student_index = whole_marks[whole_marks['Students']==names[student]].index[0].astype(int)
                                for data_students in range(np.shape(data)[0]):
                                    surname = data.loc[data_students, "Apellido(s)"]
                                    name = data.loc[data_students, "Nombre"]
                                    complete_name = str(surname) + ", " + str(name)
                                    if complete_name == names[student]:
                                        data_student_index = data_students
                                whole_marks=whole_marks.fillna("NP")
                                whole_marks.at[marks_student_index, shortname]=data.loc[data_student_index, data_cols[7]]/max_mark*10
                                efficiency=efficiency.fillna("NP")
                                efficiency.at[marks_student_index, shortname]=points_per_min[data_student_index]
                else:
                    messagebox.showinfo("Warning", "File not analyzed: " + columns_ok)
            
            pb['value']=10
            root.update_idletasks()
            
            backup_wm = whole_marks.copy()
            backup_wm = backup_wm.replace("NP", "0")
            backup_wm = backup_wm.replace("-", "0")
            backup_wm = backup_wm.stack().str.replace(',','.').unstack()
            print(backup_wm)
            cols=[i for i in backup_wm.columns if i not in ["Students"]]
            for col in cols:
                   backup_wm[col]=pd.to_numeric(backup_wm[col], downcast="float")
            whole_marks["Average marks"]=backup_wm.drop('Students', axis=1).mean(axis=1)
            whole_marks["Std. deviation"]=backup_wm.drop('Students', axis=1).std(axis=1)
            
            # Plot normal distribution of the whole marks
            plt.figure(figsize = (25, 15), dpi = 80)
            x=np.linspace(-2,14,320)
            normal_distribution=[]
            for index in range(np.shape(whole_marks)[0]):
                std_dev=whole_marks.loc[index,"Std. deviation"]
                if whole_marks.loc[index,"Std. deviation"]==0:
                    std_dev=0.0001
                normal_distribution.append((1/(std_dev*np.sqrt(2*math.pi)))*np.exp(-(x-whole_marks.loc[index,"Average marks"])**2/(2*(std_dev**2))))
                draw=plt.plot(x,normal_distribution[index],label=whole_marks.loc[index, "Students"])
            std_dev=whole_marks["Std. deviation"].mean()
            if whole_marks["Std. deviation"].mean() == 0:
                std_dev=0.0001
            normal_distribution.append((1/(std_dev*np.sqrt(2*math.pi)))*np.exp(-(x-whole_marks["Average marks"].mean())**2/(2*(std_dev**2))))
            plt.plot(x,normal_distribution[index], '--')
            plt.tick_params(axis='x', labelsize=12)
            plt.tick_params(axis='y', labelsize=20)
            plt.grid()
            plt.legend(fontsize=16)
            plt.xticks(np.linspace(-2,14,32))
            plt.xlabel('x', fontsize=24)
            plt.ylabel('F(x)', fontsize=24)
            plt.title("Course marks normal distribution", fontsize=24)
            plt.savefig(parent_dir + "Course/course_marks_normal_distribution.png")
            plt.show()
            print(whole_marks)

            n_questions=np.shape(backup_wm)[1]-1
            n_students=np.shape(backup_wm)[0]
            backup_ef = efficiency.copy()
            backup_ef = backup_ef.replace("NP", "0")
            cols=[i for i in backup_ef.columns if i not in ["Students"]]
            for col in cols:
                   backup_ef[col]=pd.to_numeric(backup_ef[col], downcast="float")
            efficiency["Average marks"]=np.round(backup_ef.drop('Students', axis=1).mean(axis=1),2)
            plt.figure(figsize = (25, 10), dpi = 80)
            plt.xticks(range(n_students),efficiency.loc[:,"Students"],rotation=90)
            print("Efficiency", efficiency)
            plt.bar(efficiency.loc[:,"Students"], np.round(efficiency.loc[:,"Average marks"], decimals=2))
            plt.title("Student's efficiency")
            plt.xlabel("Student", fontsize=14)
            plt.ylabel("Efficiency (points per minute)", fontsize=14)
            
            for index, value in enumerate(efficiency.loc[:,"Average marks"]):
                    plt.text(index,value, str(value), fontsize=14)
            plt.savefig(parent_dir + "Course/course_efficiency.png")
            plt.show()

            califications_conversion = []
            califications_conversion = whole_marks['Average marks'].apply(lambda s: ScoreMark(int(s)))
            califications_conversion.value_counts(sort = True).plot.pie(autopct='%1.2f%%', title="Students' marks", shadow=True)
            plt.ylabel("")
            plt.savefig(parent_dir + "Course/whole_course_piechart.png")
            
            pb['value']=20
            root.update_idletasks()

            # Print marks bar chart
            fig = plt.figure(figsize = (5,6))
            ax=whole_marks[['Average marks']].astype(int).value_counts(sort=False).plot.bar(rot=0)
            plt.xlabel('Integer value of the mark')
            plt.ylabel('Distribution of marks')
            plt.title("Counts of intervals for each student's mark")
            plt.grid()
            plt.savefig(parent_dir + "Course/whole_course_barchart.png")
            plt.show()

            # Print distribution of marks for each question
            question_grades=backup_wm
            question_grades=question_grades.drop("Students", axis=1)
            question_grades=question_grades.replace(",",".")
            fig = plt.figure(figsize = (15,10))
            ax = fig.gca()
            question_grades.hist(ax=ax)
            plt.suptitle("Distribution of marks for each quiz")
            plt.xlabel('Integer value of the mark')
            plt.ylabel('Distribution of marks')
            plt.savefig(parent_dir + "Course/whole_course_exams_barchart.png")
            plt.show()

            pb['value']=50
            root.update_idletasks()
            
            # Calculation of IRT
            avg_marks=whole_marks['Average marks'].to_numpy()/10
            avg_marks = np.where(avg_marks==1,0.999,avg_marks)
            print("Whole marks", whole_marks)

            theta = np.log(avg_marks/(1-avg_marks)) # level of the student
            #print("Theta", theta)

            print("Means", backup_wm.mean())
            questions_means=backup_wm.mean().to_numpy()/10
            questions_means=np.where(questions_means==1,0.999,questions_means)
            print(questions_means)
            beta = np.log((1-questions_means)/questions_means) # difficulty of the question
            for it in range(len(beta)):
                if(beta[it] > 3):
                    beta[it] = 3
                if(beta[it] < -3):
                    beta[it] = -3
                if(beta[it] == math.inf):
                    beta[it] = 3
                if(beta[it] == -math.inf):
                    beta[it] == -3
            #print("Beta", beta)

            probability=[]
            for i in range(len(theta)):
                    probability.append([])
                    for j in range(len(beta)):
                        probability[i].append(np.exp(theta[i]-beta[j])/(1+np.exp(theta[i]-beta[j]))) # probability of answering correctly
            print(probability)

            # Print normalized beta
            division=np.max(beta) - np.min(beta)
            if division==0:
                division=0.0001
            norm_beta = 10*(beta - np.min(beta)) / division
            for it in range(len(norm_beta)):
                if norm_beta[it] == math.inf:
                    norm_beta[it]=3
            norm_beta = np.round(norm_beta, decimals=2)
            plt.figure(figsize = (25, 10), dpi = 80)
            mpl.rcParams["font.size"] = 24
            plt.xticks(range(n_questions),range(1,n_questions+1))
            #for index, value in enumerate(norm_beta):
            #        plt.text(index,value, str(value))
            plt.xlabel('Exam', fontsize=24)
            plt.ylabel('Difficulty', fontsize=24)
            plt.bar(range(len(norm_beta)),norm_beta)
            plt.savefig(parent_dir + "Course/dif_questions.png")
            plt.show()

            # Print probabilities
            probability=np.array(probability).reshape(n_students,n_questions)
            #print(probability)
            plt.figure(figsize = (50, 30), dpi = 80)
            legend_question = []
            #legend_question = "Question" + range(0,n_questions)
            legend_question = ["{} {}".format("Question", i) for i in range(1, n_questions+1)]
            plt.plot(probability, linewidth=3)
            plt.grid()
            plt.xticks(range(0,n_students),whole_marks.loc[:,"Students"],rotation=90)
            plt.xlabel('Student', fontsize=24)
            plt.ylabel('Probability of answering correctly', fontsize=24)
            plt.scatter(range(0,n_students),probability.mean(axis=1), label="Mean probability for each student", linewidth=10)
            plt.plot(range(0,n_students),probability.mean(axis=1), linewidth=5)
            plt.legend()
            plt.savefig(parent_dir + "Course/course_probability_IRT.png")
            plt.show()
            plt.close()

            # Print IRT curves
            mpl.rcParams["font.size"] = 10
            theta = np.log(avg_marks/(1-avg_marks))/10 # level of the student
            #print(theta)

            questions_means=backup_wm.mean().to_numpy()/10
            #print(questions_means)
            beta = np.log((1-questions_means)/questions_means) # difficulty of the question
            #print("Beta", beta)

            probability=[]
            theta = np.linspace(-3, 3, num=100)
            mean_beta = np.mean(beta)
            for i in range(len(beta)):
                    probability.append([])
                    for j in range(len(theta)):
                          probability[i].append(np.exp(theta[j]-beta[i])/(1+np.exp(theta[j]-beta[i]))) # probability of answering correctly
            #print(probability)

            # Print probabilities
            probability=np.array(probability).reshape(len(beta),len(theta))
            #print(probability)

            # Create directory for ICCs
            dirName = parent_dir + 'Course/ICC'
            try:
                    # Create target Directory
                    os.mkdir(dirName)
                    print("Directory " , dirName ,  " created ") 
            except FileExistsError:
                    print("Directory " , dirName ,  " already exists")

            # Create ICCs    
            for k in range(0,np.shape(probability)[0]):
                    plt.figure(figsize = (5, 4), dpi = 80)
                    plt.plot(theta,probability[k,:])
                    plt.grid()
                    #plt.xticks(range(0,n_students),names,rotation=90)
                    plt.xlabel('\u03B8', fontsize=12)
                    plt.ylabel('P(\u03B8)', fontsize=12)
                    #plt.scatter(theta,probability.mean(axis=1), label="Mean probability for each student")
                    #plt.plot(range(0,n_students),probability.mean(axis=1))
                    #plt.legend()
                    plt.title("ICC for Exam {}" .format(k+1))
                    plt.savefig(dirName + "/ICC_Exam{:03}.png".format(k+1))
                    plt.show()
                    plt.close()

            pb['value']=70
            root.update_idletasks()
            course_pdf(parent_dir + "Course/", pdf, pb)
        else:
            shortname, sep, tail = (os.path.basename(filenames[0])).partition('.')
            dirName = 'Quiz_' + shortname
            data = pd.ExcelFile(filenames[0]).parse()
            data.drop(data.tail(1).index,inplace=True)
            columns_ok=check_columns(data, qr_cols)
            print(columns_ok)
            if columns_ok == "All columns are available in the dataframe":
                path = dirName
                try:
                    os.mkdir(path)
                    print("Directory " + path + " created")
                except FileExistsError:
                    print("Directory " + path + " already exists")
                # Read names of students
                data = pd.ExcelFile(filenames[0]).parse()
                data.drop(data.tail(1).index,inplace=True)
                names=[]
                names = data[['Apellido(s)', 'Nombre']].apply(lambda x: ', '.join(x[x.notnull()]), axis = 1)
                SingleQuizReport(path,shortname,filename,pdf,names)
            else:
                messagebox.showinfo("Warning", "File not analyzed: " + columns_ok)
        pb['value']=100
        lab.destroy()
        pb.grid_forget()
        root.update_idletasks()
        root.update()
        pdf.output(asksaveasfilename(title="Save output file", defaultextension=".pdf", filetypes=(("PDF File", "*.pdf"),("All Files", "*.*"))), "F")
        root.update()
    except (FileNotFoundError,IndexError):
        pb['value']=100
        lab.destroy()
        pb.grid_forget()
        root.update_idletasks()
        messagebox.showinfo("Warning", "No file has been chosen!")
    except ValueError:
        pb['value']=100
        lab.destroy()
        pb.grid_forget()
        root.update_idletasks()
        messagebox.showinfo("Warning", "No Excel file has been chosen!")
        
def openLogReport():
    
    try:
        root.update()
        log_file = filedialog.askopenfilename(title="Choose a Log Report",filetypes=(("Excel Files (.xlsx)","*.xlsx"),("All Files","*.*")))
        root.update()
        logs = pd.read_excel(log_file)
        logs_cols = ['Hora', 'Nombre completo del usuario', 'Usuario afectado', 'Contexto del evento', 'Componente', 'Nombre evento', 'Descripción', 'Origen', 'Dirección IP']
        columns_ok = check_columns(logs,logs_cols)
        
        pb = ttk.Progressbar(root, length=100, mode='determinate')
        pb.grid(column=0, padx=200, sticky="w")
        lab = tk.Label(root, text='Calculating...')
        lab.grid(column=0, padx=210, sticky="w")
            
        if columns_ok == "All columns are available in the dataframe.":
            # OBTENTION OF ENROLLED STUDENTS
            enrolled_students = []
            enrollment_date = []
            for i in reversed(range(0,len(logs))):
              if (logs.iloc[i]["Nombre completo del usuario"] not in enrolled_students and logs.iloc[i]["Nombre completo del usuario"] != "traspaso" and logs.iloc[i]["Nombre completo del usuario"] != "-" and logs.iloc[i]["Nombre completo del usuario"] != "Admin Aula Global" and "como" not in logs.iloc[i]["Nombre completo del usuario"]):
                enrolled_students.append(logs.iloc[i]["Nombre completo del usuario"])
                date = datetime.strptime(logs.iloc[i]["Hora"], '%d/%m/%Y %H:%M')
                enrollment_date.append(str(date.day) + "/" + str(date.month) + "/" + str(date.year))
              if (logs.iloc[i]["Componente"]=="Sistema" and logs.iloc[i]["Nombre evento"]=="Usuario desmatriculado del curso"):
                enrolled_students.remove(logs.iloc[i]["Usuario afectado"])
                enrollment_date.remove(str(date.day) + "/" + str(date.month) + "/" + str(date.year))

            pb['value']=30
            root.update_idletasks()
            
            quiz_status = pd.DataFrame(enrolled_students, columns=["Students"]) # Status of each student for each quiz
            start_IP = pd.DataFrame(enrolled_students, columns=["Students"]) # Used IP to start a quiz
            finish_IP = pd.DataFrame(enrolled_students, columns=["Students"]) # Used IP to send the attempt of the quiz
            start_date = pd.DataFrame(enrolled_students, columns=["Students"])
            duration = pd.DataFrame(enrolled_students, columns=["Students"]) # Duration in minutes for the quiz
            consults = pd.DataFrame(enrolled_students, columns=["Students"]) # Consults interactions
            forum = pd.DataFrame(enrolled_students, columns=["Students"]) # Forum interactions
            polls = pd.DataFrame(enrolled_students, columns=["Students"]) # Polls interactions
            resources = pd.DataFrame(enrolled_students, columns=["Students"]) # Resources interactions
            tasks = pd.DataFrame(enrolled_students, columns=["Students"]) # Tasks interactions
            level = pd.DataFrame(enrolled_students, columns=["Students"]) # Level interactions
            quiz_attempts = pd.DataFrame(enrolled_students, columns=["Students"]) # Quiz attempts for each student in each quiz
            modules = pd.DataFrame(enrolled_students, columns=["Students"]) # Modules accesses by each student
            activity = pd.DataFrame(enrolled_students, columns=["Students"]) # Interactions per day by each student            
            
            level["Level"]="1"
            forum["Created threads"]="0"
            forum["Sent messages"]="0"
            forum["Forum interactions"]="0"
            polls["Polls interactions"]="0"
            resources["Resources interactions"]="0"
            
            # MOVING THROUGH THE LOGS
            for i in reversed(range(len(logs))):
                try:
                  active_user = logs.iloc[i]["Nombre completo del usuario"]
                  if active_user in enrolled_students:
                     active_user_index = quiz_status[quiz_status['Students']==active_user].index[0].astype(int)
                  affected_user = logs.iloc[i]["Usuario afectado"]
                  if affected_user in enrolled_students:
                     affected_user_index = quiz_status[quiz_status['Students']==affected_user].index[0].astype(int)
                  component = logs.iloc[i]["Componente"]
                  event_name = logs.iloc[i]["Nombre evento"]
                  description = logs.iloc[i]["Descripción"]
                  used_IP = logs.iloc[i]["Dirección IP"]
                  date = datetime.strptime(logs.iloc[i]["Hora"], '%d/%m/%Y %H:%M')
                  context = logs.iloc[i]["Contexto del evento"]

                  todays_date = str(date.day) + "/" + str(date.month) + "/" + str(date.year)
                  if todays_date not in activity.columns and active_user!="traspaso":
                    activity[todays_date]="0"
                  elif active_user!="traspaso":
                    activity.at[active_user_index,todays_date]=str(int(activity.loc[active_user_index,todays_date])+1)

                  # Getting the 'course module id'   
                  #print(component,event_name,description)
                  #if (component == "Wooclap" and event_name == "Módulo de curso visto") or (component == "Tarea" and event_name != "Instancia del módulo del curso visualizada") or component == "Módulo de encuesta" or (component == "Herramienta externa" and event_name!="Instancia del módulo del curso visualizada") or component == "Entregas de texto en línea" or component == "Carpeta" or component == "URL" or (component == "Foro" and event_name != "Instancia del módulo del curso visualizada") or (component == "Cuestionario" and event_name!="Excepción de cuestionario creada") or component == "Recurso" or (component == "Sistema" and (event_name == "Módulo de curso creado" or event_name == "Módulo de curso actualizado" or event_name == "Módulo de curso eliminado")):
                  #  res = re.search("with course module id '(.*)'.", description)
                  #  id = int(res.group(1))
                  #elif (component == "Sistema" and event_name == "Usuario matriculado en el curso") or (component == "Foro" and event_name == "Instancia del módulo del curso visualizada") or (component == "Tarea" and event_name == "Instancia del módulo del curso visualizada"):
                  #  res = re.search("in the course with id '(.*)'.", description)
                  #  id = int(res.group(1))
                  #elif component == "Sistema" and event_name == "Rol asignado":
                  #  res = re.search("to the user with id '(.*)'.", description)
                  #  id = int(res.group(1))
                  #elif component == "Sistema" and event_name == "Curso visto":
                  #  res = re.search("viewed the course with id '(.*)'.", description)
                  #  id = int(res.group(1))
                  #print(id)

                  # COMPONENTS MANAGEMENT
                  if component == "¡Sube de nivel!":
                    level.at[active_user_index, "Level"]=str(int(level.loc[active_user_index, "Level"])+1)
                    
                  if component == "Consulta":
                    if context not in consults and (not context.startswith("Curso") and not context.startswith("Otro")):
                        consults[context]="0"
                    if event_name == "Consulta respondida":
                        consults.at[active_user_index,context]="1"
                        
                  if component == "Cuestionario":
                    if context not in quiz_status.columns and not context.startswith("Curso"):
                        quiz_status[context]="UNOPENED"
                        start_IP[context]=""
                        finish_IP[context]=""
                        start_date[context]=date
                        duration[context]=""
                        quiz_attempts[context]="0"
                    if not context.startswith("Curso"):
                        if event_name == "Ha comenzado el intento":
                            quiz_status.at[active_user_index, context]="OPENED"
                            start_IP.at[active_user_index, context]=used_IP
                            start_date.at[active_user_index, context]=date
                            quiz_attempts.at[active_user_index, context]=str(int(quiz_attempts.loc[active_user_index, context])+1)
                        if event_name == "Intento enviado":
                            if quiz_status.loc[active_user_index, context] == "CHEATED (Opened resources while answering)" and used_IP == start_IP.loc[active_user_index, context]:
                              finish_IP.at[active_user_index, context]=used_IP
                              duration.at[active_user_index, context]=((date-start_date.loc[active_user_index, context]).total_seconds())/60
                            elif quiz_status.loc[active_user_index, context] == "CHEATED (Opened resources while answering)" and used_IP != start_IP.loc[active_user_index, context]:
                              quiz_status.at[active_user_index, context]="CHEATED (Opened resources while answering and used a different device for sending the attempt)"
                              finish_IP.at[active_user_index, context]=used_IP
                              duration.at[active_user_index, context]=((date-start_date.loc[active_user_index, context]).total_seconds())/60
                            elif quiz_status.loc[active_user_index, context] == "OPENED" and used_IP != start_IP.loc[active_user_index, context]:
                              quiz_status.at[active_user_index, context]="CHEATED (Used a different device for sending the attempt)"
                              finish_IP.loc[active_user_index, context]=used_IP
                              duration.at[active_user_index, context]=((date-start_date.loc[active_user_index, context]).total_seconds())/60
                            elif quiz_status.loc[active_user_index, context] == "OPENED" and used_IP == start_IP.loc[active_user_index, context]:
                              quiz_status.at[active_user_index, context]="ATTEMPT SENT"
                              finish_IP.loc[active_user_index, context]=used_IP
                              duration.at[active_user_index, context]=((date-start_date.loc[active_user_index, context]).total_seconds())/60
                              #print(type(duration.loc[active_user_index, context]),type(date), type(start_date.loc[active_user_index, context]), date, start_date.loc[active_user_index, context], duration.loc[active_user_index, context])
                  if component == "Foro":
                    forum.at[active_user_index, "Forum interactions"]=str(int(forum.loc[active_user_index, "Forum interactions"])+1)
                    if event_name == "Tema creado":
                        forum.at[active_user_index, "Created threads"]=str(int(forum.loc[active_user_index, "Created threads"])+1)
                    elif event_name == "Mensaje creado":
                        forum.at[active_user_index, "Sent messages"]=str(int(forum.loc[active_user_index, "Sent messages"])+1)

                  if component == "Módulo de encuesta":
                    polls.at[active_user_index, "Polls interactions"]=str(int(polls.loc[active_user_index, "Polls interactions"])+1)

                  if component == "Recurso":
                    resources.at[active_user_index, "Resources interactions"]=str(int(resources.loc[active_user_index, "Resources interactions"])+1)                      

                  if component == "Sistema":
                    if event_name == "Módulo de curso creado" and (not context.startswith("Otro") and not context.startswith("Etiqueta")):
                        modules[context]="0"
              
                  if component == "Tarea":
                    if context not in tasks.columns and (not context.startswith("Curso") and not context.startswith("Otro")):
                            tasks[context]="0"
                    if event_name == "Se ha enviado una entrega":
                        tasks.at[active_user_index,context]=str(int(tasks.loc[active_user_index, context])+1)

                  if event_name == "Módulo de curso visto" and not context.startswith("Etiqueta") and component != "Cuestionario":
                    modules.at[active_user_index,context]="1"
                    quiz_status.loc[active_user_index,:].replace("OPENED", "CHEATED (Opened resources while answering)", inplace=True)
                    
                except Exception as e:
                    print(e)
                    continue
            
            pb['value']=80
            root.update_idletasks()

            # Get percentage of answered consults
            cols=[i for i in consults.columns if i not in ["Students"]]
            for col in cols:
                consults[col]=pd.to_numeric(consults[col], downcast="float")
            consults['Sum']=consults.drop('Students', axis=1).sum(axis=1)
            consults['Percentage']=consults['Sum']/(np.shape(consults)[1]-2)*100
            print(consults)
            
            # Get percentage of sent tasks
            cols=[i for i in tasks.columns if i not in ["Students"]]
            for col in cols:
                tasks[col]=pd.to_numeric(tasks[col], downcast="float")
            tasks['Sum']=tasks.drop('Students', axis=1).sum(axis=1)
            tasks['Percentage']=tasks['Sum']/(np.shape(tasks)[1]-2)*100
            
            # Getting average duration of students
            cols=[i for i in duration.columns if i not in ["Students"]]
            for col in cols:
                duration[col]=pd.to_numeric(duration[col], downcast="float")
            duration['Average']=duration.drop('Students', axis=1).mean(axis=1)
            
            # Getting seen modules
            cols=[i for i in modules.columns if i not in ["Students"]]
            for col in cols:
                modules[col]=pd.to_numeric(modules[col], downcast="float")
            modules['Sum']=modules.drop('Students', axis=1).sum(axis=1)
            modules['Percentage']=modules['Sum']/(np.shape(modules)[1]-2)*100

            # Getting done and cheated exams
            quiz_status = quiz_status.loc[:,~(quiz_status.astype(str)=='UNOPENED').all()]
            quiz_status['Done exams']=quiz_status.apply(lambda row: sum(row[1:]!="UNOPENED") ,axis=1)
            quiz_status['Percentage of done exams']=quiz_status['Done exams']/(np.shape(quiz_status)[1]-2)*100
            quiz_status['Cheated exams']= quiz_status.isin(["CHEATED (Opened resources while answering)", "CHEATED (Used a different device for sending the attempt)", "CHEATED (Opened resources while answering and used a different device for sending the attempt)"]).sum(1)
            quiz_status['Percentage of cheated exams']=quiz_status['Cheated exams']/(np.shape(quiz_status)[1]-4)*100            
            
            #print(quiz_status, duration, quiz_attempts,modules)

            # Calculation of average number of attempts
            backup_qa = quiz_attempts.copy()
            cols=[i for i in backup_qa.columns if i not in ["Students"]]
            for col in cols:
                    backup_qa[col]=pd.to_numeric(backup_qa[col])
            quiz_attempts['Average']=backup_qa.drop('Students', axis=1).mean(axis=1)
            print(quiz_attempts)

            # Get students profile
            backup_ac = activity.copy()
            cols=[i for i in backup_ac.columns if i not in ["Students"]]
            for col in cols:
                    backup_ac[col]=pd.to_numeric(backup_ac[col])
            n_days=np.zeros((len(enrolled_students)))            
            interactions=np.zeros((len(enrolled_students)))
            for student in range(len(enrolled_students)):
                n_days[student] = (date - datetime.strptime(enrollment_date[student], '%d/%m/%Y')).days
                act_row=backup_ac.drop(columns=["Students"])
                print((act_row.loc[student,:]>0).astype(int).sum(),n_days[student])
                interactions[student]=(act_row.loc[student,:]>0).astype(int).sum()/n_days[student]*100
            activity["Average"]=backup_ac.drop('Students', axis=1).mean(axis=1)
            activity["Active days (%)"]=interactions
            students_profile = pd.DataFrame({"Students": enrolled_students, 'Enrollment date': enrollment_date, 'Level': level["Level"], 'Sent tasks (%)': tasks["Percentage"], 'Answered consults (%)': consults["Percentage"], 'Forum interactions': forum["Forum interactions"], 'Polls interactions': polls["Polls interactions"], "Resources visualizations": resources["Resources interactions"], "Visualized modules (%)": modules["Percentage"], "Days of activity (%)": interactions, "Done exams (%)": quiz_status["Percentage of done exams"], "Cheated exams(%)": quiz_status['Percentage of cheated exams']})
            print(students_profile)

            # Calculate effort bonification
            effort_bonification=pd.DataFrame(enrolled_students, columns=["Students"])
            effort_bonification["Result"]="0"

            for student in range(0, np.shape(students_profile)[0]):
              if int(students_profile.loc[student, "Level"])>=int(pd.to_numeric(students_profile["Level"]).mean()) and float(students_profile.loc[student, "Visualized modules (%)"])>=80 and float(students_profile.loc[student, "Days of activity (%)"])>=80 and int(students_profile.loc[student, "Done exams (%)"])==100 and float(students_profile.loc[student, 'Cheated exams(%)'])==0:
                  effort_bonification.at[student, "Result"]="1"
            #print(effort_bonification)

            root.update()
            writer = filedialog.asksaveasfilename(title="Save output file", defaultextension=".xlsx")
            root.update()
            
            with pd.ExcelWriter(writer) as writer1:
                # Write each dataframe to a different worksheet.
                students_profile.to_excel(writer1, sheet_name='Student profiles')
                effort_bonification.to_excel(writer1, sheet_name="Effort bonification")
                quiz_status.to_excel(writer1, sheet_name='Quiz status')
                duration.to_excel(writer1, sheet_name='Quiz duration')
                quiz_attempts.to_excel(writer1, sheet_name='Quiz attempts')
                start_IP.to_excel(writer1, sheet_name='Quiz start IP address')
                finish_IP.to_excel(writer1, sheet_name='Quiz finish IP address')
                tasks.to_excel(writer1, sheet_name='Sent tasks')
                consults.to_excel(writer1, sheet_name='Answered consults')
                modules.to_excel(writer1, sheet_name="Visualized modules")
                activity.to_excel(writer1, sheet_name="Course activity")
                forum.to_excel(writer1, sheet_name="Forum interactions")
                # Close the Pandas Excel writer and output the Excel file.
                writer1.save()
            pb['value']=100
            pb.grid_forget()
            lab.destroy()
            root.update_idletasks()
        else:
            pb['value']=100
            pb.grid_forget()
            lab.destroy()
            root.update_idletasks()
            messagebox.showinfo("Warning", "File not analyzed: " + columns_ok)
            
    except (FileNotFoundError,IndexError):
        pb = ttk.Progressbar(root, length=100, mode='determinate')
        pb['value']=100
        pb.grid_forget()
        lab = tk.Label(root, text='Calculating...')
        lab.grid(column=0, padx=210, sticky="w")
        lab.destroy()
        root.update_idletasks()
        messagebox.showinfo("Warning", "No file has been chosen!")
    except ValueError:
        pb = ttk.Progressbar(root, length=100, mode='determinate')
        pb['value']=100
        pb.grid_forget()
        lab = tk.Label(root, text='Calculating...')
        lab.grid(column=0, padx=210, sticky="w")
        lab.destroy()
        root.update_idletasks()
        messagebox.showinfo("Warning", "No Excel file has been chosen!")
        
def openGradeReport():
    
    try:
        root.update()
        grades_file = filedialog.askopenfilename(title="Choose a Grade Report",filetypes=(("Excel Files (.xlsx)","*.xlsx"),("All Files","*.*")))
        root.update()
        grades = pd.read_excel(grades_file)
        grades_cols = ['Nombre', 'Apellido(s)', 'Dirección de correo', 'Total del curso (Real)', 'Última descarga de este curso']
        columns_ok = check_columns(grades, grades_cols)

        if columns_ok == "All columns are available in the dataframe.":
            backup_g = grades.copy()
            backup_g = backup_g.drop(columns=["Última descarga de este curso"])
            backup_g = backup_g.stack().str.replace(',','.').unstack()
            cols=[i for i in backup_g.columns if i not in ["Nombre", "Apellido(s)", "Dirección de correo"]]
            for col in cols:
                    backup_g[col]=pd.to_numeric(backup_g[col], downcast="float")  
            print(backup_g)
            reported_students = []
            reported_students = grades[['Apellido(s)', 'Nombre']].apply(lambda x: ', '.join(x[x.notnull()]), axis = 1)
            marks = pd.DataFrame(reported_students, columns=["Students"])
            marks["Quiz average"]=backup_g.filter(regex='Cuestionario:').mean(axis=1)
            marks["Quiz std. deviation"]=backup_g.filter(regex='Cuestionario:').std(axis=1)
            marks["Tasks average"]=backup_g.filter(regex='Tarea:').mean(axis=1)
            marks["Tasks std. deviation"]=backup_g.filter(regex='Tarea:').std(axis=1)
            marks["External tools average"]=backup_g.filter(regex='Herramienta externa:').mean(axis=1)
            marks["External tools std. deviation"]=backup_g.filter(regex='Herramienta externa:').std(axis=1)
            marks["Kaltura Media average"]=backup_g.filter(regex='Kaltura Media Assignment:').mean(axis=1)
            marks["Kaltura Media std. deviation"]=backup_g.filter(regex='Kaltura Media Assignment:').std(axis=1)
            marks["Marks average"]=backup_g.mean(axis=1)
            marks["Marks std. deviation"]=backup_g.std(axis=1)
            print(marks)

            shortname, sep, tail = (os.path.basename(grades_file)).partition('.')
            dirName = 'GradeReport_' + shortname
            # Create directory for each quiz
            try:
                os.mkdir("Course")
                print("Directory 'Course' created")
            except FileExistsError:
                print("Directory 'Course' already exists")    
            path = "Course/" + dirName
            # Create directory for each quiz
            try:
                os.mkdir(path)
                print("Directory " + path + " created")
            except FileExistsError:
                print("Directory " + path + " already exists")      

            # Plot normal distribution of the marks
            plt.figure(figsize = (25, 10), dpi = 80)
            x=np.linspace(-2,10,240)
            normal_distribution=[]
            average=marks["Marks average"]
            std_dev=marks["Marks std. deviation"]
            for index in range(np.shape(marks)[0]):
                normal_distribution.append((1/(marks.loc[index,"Marks std. deviation"]*np.sqrt(2*math.pi)))*np.exp(-(x-marks.loc[index,"Marks average"])**2/(2*(marks.loc[index, "Marks std. deviation"]**2))))
            normal_distribution.append((1/(marks["Marks std. deviation"].mean()*np.sqrt(2*math.pi)))*np.exp(-(x-marks["Marks average"].mean())**2/(2*(marks["Marks std. deviation"].mean()**2))))
            plt.plot(x,normal_distribution[index+1], label="Mean curve")
            plt.tick_params(axis='x', labelsize=12)
            plt.tick_params(axis='y', labelsize=20)
            plt.grid()
            plt.legend(fontsize=12)
            plt.xticks(np.linspace(0,10,24))
            plt.xlabel('x', fontsize=18)
            plt.ylabel('F(x)', fontsize=18)
            plt.title("Marks normal distribution", fontsize=24)
            plt.savefig(path + "marks_normal_distribution.png")
            plt.show()    
            plt.close()
            
            # Plot quiz normal distribution
            plt.figure(figsize = (25, 10), dpi = 80)
            normal_distribution=[]
            for index in range(np.shape(marks)[0]):
                normal_distribution.append((1/(marks.loc[index,"Quiz std. deviation"]*np.sqrt(2*math.pi)))*np.exp(-(x-marks.loc[index,"Quiz average"])**2/(2*(marks.loc[index,"Quiz std. deviation"]**2))))
            normal_distribution.append((1/(marks["Quiz std. deviation"].mean()*np.sqrt(2*math.pi)))*np.exp(-(x-marks["Quiz average"].mean())**2/(2*(marks["Quiz std. deviation"].mean()**2))))
            plt.plot(x,normal_distribution[index+1], label="Mean curve")
            plt.tick_params(axis='x', labelsize=12)
            plt.tick_params(axis='y', labelsize=20)
            plt.grid()
            plt.legend(fontsize=12)
            plt.xticks(np.linspace(0,10,24))
            plt.xlabel('x', fontsize=18)
            plt.ylabel('F(x)', fontsize=18)
            plt.title("Quiz normal distribution", fontsize=24)
            plt.savefig(path + "quiz_normal_distribution.png")
            plt.show() 
            plt.close()
            
            # Tasks normal distribution
            plt.figure(figsize = (25, 10), dpi = 80)
            normal_distribution=[]
            for index in range(np.shape(marks)[0]):
                normal_distribution.append((1/(marks.loc[index,"Tasks std. deviation"]*np.sqrt(2*math.pi)))*np.exp(-(x-marks.loc[index,"Tasks average"])**2/(2*(marks.loc[index,"Tasks std. deviation"]**2))))
            normal_distribution.append((1/(marks["Tasks std. deviation"].mean()*np.sqrt(2*math.pi)))*np.exp(-(x-marks["Tasks average"].mean())**2/(2*(marks["Tasks std. deviation"].mean()**2))))
            plt.plot(x,normal_distribution[index+1], label="Mean curve")
            plt.tick_params(axis='x', labelsize=12)
            plt.tick_params(axis='y', labelsize=20)
            plt.grid()
            plt.legend(fontsize=12)
            plt.xticks(np.linspace(0,10,24))
            plt.xlabel('x', fontsize=18)
            plt.ylabel('F(x)', fontsize=18)
            plt.title("Tasks normal distribution", fontsize=24)
            plt.savefig(path + "tasks_normal_distribution.png")
            plt.show()  
            plt.close()
            
            # External tools normal distribution
            plt.figure(figsize = (25, 10), dpi = 80)
            normal_distribution=[]
            for index in range(np.shape(marks)[0]):
                normal_distribution.append((1/(marks.loc[index,"External tools std. deviation"]*np.sqrt(2*math.pi)))*np.exp(-(x-marks.loc[index,"External tools average"])**2/(2*(marks.loc[index,"External tools std. deviation"]**2))))
            normal_distribution.append((1/(marks["External tools std. deviation"].mean()*np.sqrt(2*math.pi)))*np.exp(-(x-marks["External tools average"].mean())**2/(2*(marks["External tools std. deviation"].mean()**2))))
            plt.plot(x,normal_distribution[index+1], label="Mean curve")
            plt.tick_params(axis='x', labelsize=12)
            plt.tick_params(axis='y', labelsize=20)
            plt.grid()
            plt.legend(fontsize=12)
            plt.xticks(np.linspace(0,10,24))
            plt.xlabel('x', fontsize=18)
            plt.ylabel('F(x)', fontsize=18)
            plt.title("External tools normal distribution", fontsize=24)
            plt.savefig(path + "external_tools_normal_distribution.png")
            plt.show() 
            plt.close()
            
            # Kaltura Media normal distribution
            plt.figure(figsize = (25, 10), dpi = 80)
            normal_distribution=[]
            for index in range(np.shape(marks)[0]):
                normal_distribution.append((1/(marks.loc[index,"Kaltura Media std. deviation"]*np.sqrt(2*math.pi)))*np.exp(-(x-marks.loc[index,"Kaltura Media average"])**2/(2*(marks.loc[index,"Kaltura Media std. deviation"]**2))))
            normal_distribution.append((1/(marks["Kaltura Media std. deviation"].mean()*np.sqrt(2*math.pi)))*np.exp(-(x-marks["Kaltura Media average"].mean())**2/(2*(marks["Kaltura Media std. deviation"].mean()**2))))
            plt.plot(x,normal_distribution[index+1], label="Mean curve")
            plt.tick_params(axis='x', labelsize=12)
            plt.tick_params(axis='y', labelsize=20)
            plt.grid()
            plt.legend(fontsize=12)
            plt.xticks(np.linspace(0,10,24))
            plt.xlabel('x', fontsize=18)
            plt.ylabel('F(x)', fontsize=18)
            plt.title("Kaltura Media normal distribution", fontsize=24)
            plt.savefig(path + "kaltura_media_normal_distribution.png")
            plt.show()
            plt.close()

            root.update()
            writer = filedialog.asksaveasfilename(title="Save output file", defaultextension=".xlsx")
            root.update()
            
            with pd.ExcelWriter(writer) as writer1:
                marks.to_excel(writer1,sheet_name="Marks")
                worksheet = writer1.sheets['Marks']
                worksheet.insert_image('N1', path + "marks_normal_distribution.png")
                worksheet.insert_image('AV1', path + "quiz_normal_distribution.png")
                worksheet.insert_image('N44', path + "external_tools_normal_distribution.png")
                worksheet.insert_image('AV44', path + "kaltura_media_normal_distribution.png")
                writer1.save()
        else:
            messagebox.showinfo("Warning", "File not analyzed: " + columns_ok)
            
    except (FileNotFoundError,IndexError):
        messagebox.showinfo("Warning", "No file has been chosen!")
    except ValueError:
        messagebox.showinfo("Warning", "No Excel file has been chosen!")
        
img = tk.PhotoImage(file ="uc3m.png")
panel = tk.Label(root, image = img)
panel.grid(row=0,column=0,sticky="w")

img0 = tk.PhotoImage(file ="Statoodle.png")
panel1 = tk.Label(root, image = img0)
panel1.grid(row=1,column=0, padx=0, pady=10, sticky="w")

img1 = tk.PhotoImage(file ="lemait.png")
panel2 = tk.Label(root, image = img1)
panel2.grid(row=1,column=0, padx=320, pady=10, sticky="w")

quiz_frame = LabelFrame(root, width=68, text="Single Quiz Report Analytics")
quiz_frame.grid(column=0, row=8, padx=10, pady=10, sticky="w")
label = Label(quiz_frame, width=68, text="Introduce a Quiz Report (.xlsx), and its stats will be obtained. The output will be a PDF file.").pack(anchor="w")
btn = Button(quiz_frame, bg='#abd7eb', text ="Import Quiz Report and Analyze", command = openSingleQuizReport).pack()

quiz_frame1 = LabelFrame(root, width=68, text="Multiple Quiz Report Analytics")
quiz_frame1.grid(column=0, padx=10, pady=10, sticky="w")
label = Label(quiz_frame1, width=68, text="Introduce multiple Quiz Reports (.xlsx). Same process as the Single Quiz Report Analytics,\nincluding a whole course resume. The output will be a PDF file.").pack()
btn1 = Button(quiz_frame1, bg='#abd7eb', text ="Import Quiz Reports and Analyze", command = openMultipleQuizReport).pack()
  
quiz_frame2 = LabelFrame(root, width=68, text="Log Report Analytics")
quiz_frame2.grid(column=0, row=14, padx=10, pady=10, sticky="w")
label = Label(quiz_frame2, width=68, text="Introduce a Log Report (.xlsx). The output will be an Excel file (.xlsx), defining a profile\nfor each student and including features such as a cheating detector and\na constancy evaluator.").pack()
btn2 = Button(quiz_frame2, bg='#abd7eb', text ="Import Log Report and Analyze", command = openLogReport).pack()

quiz_frame3 = LabelFrame(root, width=68, text="Grade Report Analytics")
quiz_frame3.grid(column=0, padx=10, pady=10, sticky="w")
label = Label(quiz_frame3, width=68, text="Introduce a Grade Report (.xlsx). The output will be an Excel file (.xlsx), defining the\naverage marks and standard deviations for each kind of task.").pack()
btn3 = Button(quiz_frame3, bg='#abd7eb', text ="Import Grade Report and Analyze", command = openGradeReport).pack()

icon = tk.PhotoImage(file = "icon.png")
root.tk.call('wm', 'iconphoto', root._w, icon)

root.mainloop()